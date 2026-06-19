"""
Shared Excel parsing utilities for FinanceAI Lab.

All modules that accept Excel uploads use these functions.
openpyxl reads the raw file. pandas structures it into a dataframe
(a table with rows and columns) for easy processing.
"""

import re
import pandas as pd
from io import BytesIO


# Keywords that appear in a real column-header row of a finance file
_HEADER_KEYWORDS = {
    'budget', 'actual', 'variance', 'item', 'description', 'category',
    'account', 'line', 'period', 'amount', 'cost', 'revenue', 'income',
    'expense', 'total', 'net', 'gross', 'sales', 'qty', 'quantity',
}

# Patterns that mark a footer / footnote row (not data)
_FOOTER_PATTERN = re.compile(
    r'^(source|note|footnote|all figures|simulated|prepared by|disclaimer)',
    re.IGNORECASE,
)


def _find_header_row(data: bytes) -> int:
    """
    Scans the first 15 rows to find the actual column-header row.
    Returns a 0-based row index suitable for pd.read_excel(header=N).

    Many real-world Excel files have a title block (company name, period,
    subtitle) above the actual column headers. This function finds the first
    row that looks like a real header: at least 3 non-empty cells AND at
    least one cell contains a finance keyword.
    """
    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
    ws = wb.active

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True)):
        non_null = [c for c in row if c is not None]
        if len(non_null) < 3:
            continue
        for cell in non_null:
            if isinstance(cell, str):
                if any(kw in cell.lower() for kw in _HEADER_KEYWORDS):
                    return row_idx  # 0-based — pandas header= counts from 0

    return 0  # fallback: treat row 1 as the header


def parse_variance_sheet(uploaded_file) -> pd.DataFrame:
    """
    Reads a budget vs. actual Excel file and returns a clean dataframe.

    Handles files where a title block (company name, subtitles) sits above
    the real column headers — a very common pattern in finance exports.
    Also strips footer / footnote rows at the bottom.

    Returns the raw dataframe. Modules that display it should call
    build_display_df() to get a formatted copy for the screen.
    """
    data = uploaded_file.read()
    uploaded_file.seek(0)   # reset so the file can be read again if needed

    header_row = _find_header_row(data)
    df = pd.read_excel(BytesIO(data), sheet_name=0, header=header_row)
    df.columns = [str(c).strip() for c in df.columns]
    df = df.dropna(how="all")  # drop completely empty rows

    # Strip footer rows: first column contains a long explanatory note
    first_col = df.columns[0]
    is_footer = df[first_col].astype(str).str.match(_FOOTER_PATTERN, na=False)
    df = df[~is_footer].reset_index(drop=True)

    return df


def parse_generic_sheet(uploaded_file, sheet_index: int = 0) -> pd.DataFrame:
    """
    Generic parser for any single-sheet Excel file.
    Returns a clean dataframe with whitespace stripped from column names.
    """
    df = pd.read_excel(BytesIO(uploaded_file.read()), sheet_name=sheet_index)
    df.columns = [str(c).strip() for c in df.columns]
    df = df.dropna(how="all")
    return df


def extract_file_metadata(uploaded_file) -> dict:
    """
    Reads the Excel file and tries to extract company name and reporting period
    from the content — sheet name, title cell (A1), or any text in the first
    three rows. Returns a dict with 'company' and 'period' keys.
    Falls back to empty strings if nothing is found.
    """
    import re
    from io import BytesIO
    import openpyxl

    result = {"company": "", "period": ""}

    try:
        data = uploaded_file.read()
        uploaded_file.seek(0)  # reset so the file can be read again by the parser
        wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
        ws = wb.active

        # Collect text candidates: sheet name + first 3 rows of column A
        candidates = [ws.title or ""]
        for row in ws.iter_rows(min_row=1, max_row=3, max_col=3, values_only=True):
            for cell in row:
                if cell and isinstance(cell, str):
                    candidates.append(cell)

        full_text = " | ".join(candidates)

        # Period patterns: Q1 2026, H1 2026, FY2026, Jan 2026, January 2026, YTD 2026
        period_match = re.search(
            r'(Q[1-4]\s*\d{4}|H[1-2]\s*\d{4}|FY\s*\d{4}|YTD\s*\d{4}|'
            r'(?:Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|'
            r'Jul(?:y)?|Aug(?:ust)?|Sep(?:tember)?|Oct(?:ober)?|Nov(?:ember)?|'
            r'Dec(?:ember)?)\s*\d{4})',
            full_text, re.IGNORECASE
        )
        if period_match:
            result["period"] = period_match.group(0).strip()

        # Company name: look for text before a dash or "Budget" or "Variance" in A1
        if candidates[1:]:  # first content row
            first_cell = candidates[1]
            # Strip period text and common finance keywords to isolate company name
            company_text = re.sub(
                r'(Q[1-4]\s*\d{4}|H[1-2]\s*\d{4}|FY\s*\d{4}|YTD\s*\d{4}|'
                r'budget|actual|variance|vs\.?|all figures.*|simulated.*)',
                '', first_cell, flags=re.IGNORECASE
            )
            # Split on common separators and take the first meaningful chunk
            for sep in ['—', '-', '|', ':']:
                if sep in company_text:
                    company_text = company_text.split(sep)[0]
                    break
            company_text = company_text.strip().strip('—-|:').strip()
            if len(company_text) > 2:
                result["company"] = company_text

    except Exception:
        pass  # if anything fails, caller uses fallback values

    return result


def _to_number(val) -> float | None:
    """Converts a cell value to float. Returns None if not numeric."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).replace(",", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return None


def _find_columns(header_row: tuple) -> dict:
    """
    Scans a header row and returns column indices for Budget, Actual,
    Variance, and Variance %. Returns None for any not found.
    """
    BUDGET_KW  = ["budget", "plan", "target", "bud"]
    ACTUAL_KW  = ["actual", "actuals", "real", "result"]
    VARPCT_KW  = ["% var", "var %", "variance %", "% variance", "pct var",
                  "var to budget", "var to plan", "% to budget"]
    VAR_KW     = ["variance", "var $", "var(", "diff"]

    cols = {"budget": None, "actual": None, "variance_pct": None, "variance": None}

    for i, cell in enumerate(header_row):
        if cell is None:
            continue
        low = str(cell).lower().strip()
        if cols["budget"]      is None and any(k in low for k in BUDGET_KW):
            cols["budget"] = i
        elif cols["actual"]    is None and any(k in low for k in ACTUAL_KW):
            cols["actual"] = i
        elif cols["variance_pct"] is None and any(k in low for k in VARPCT_KW):
            cols["variance_pct"] = i
        elif cols["variance"]  is None and any(k in low for k in VAR_KW):
            cols["variance"] = i

    return cols


def parse_variance_sheet_robust(file_bytes: bytes) -> pd.DataFrame:
    """
    Robust cell-by-cell parser for complex finance Excel files using openpyxl.

    Handles real-world patterns that pandas cannot manage on its own:
    - Multi-column labels (line item names spread across columns D/E/F by indentation)
    - Sparse header rows (empty cells between column names)
    - Section header rows mixed into data (e.g. '*FACILITY EXPENSES' with no numbers)
    - Merged title blocks above the real column headers
    - Leading asterisks in cell values (*Rent, *Utilities)
    - Extra comparison columns (Prior Year, Prior Period, % of Rev)

    Returns a clean 4-column dataframe: Line Item, Budget, Actual, Variance %
    Raises ValueError with a clear message if the file cannot be parsed.
    """
    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        raise ValueError("The file appears to be empty.")

    # Find the header row and the column positions for key fields
    header_idx = _find_header_row(file_bytes)
    header_row = all_rows[header_idx]
    col_idx = _find_columns(header_row)

    if col_idx["budget"] is None or col_idx["actual"] is None:
        raise ValueError(
            "Could not find Budget and Actual columns. "
            "Ensure the file has columns named 'Budget' (or 'Plan') "
            "and 'Actual' (or 'Actuals')."
        )

    records = []
    for row in all_rows[header_idx + 1:]:
        if not row or all(c is None for c in row):
            continue

        # Read numeric values — skip rows where both are missing (section headers)
        budget_val = _to_number(row[col_idx["budget"]] if col_idx["budget"] < len(row) else None)
        actual_val = _to_number(row[col_idx["actual"]] if col_idx["actual"] < len(row) else None)
        if budget_val is None and actual_val is None:
            continue

        # Find the label: scan left to right, take the first non-empty text cell
        # This handles labels in column D, E, or F depending on indentation level
        label = None
        for cell in row:
            if cell is None:
                continue
            cell_str = str(cell).strip()
            if not cell_str or cell_str.lower() in ("none", "nan"):
                continue
            # Skip cells that are purely numeric
            try:
                float(cell_str.replace(",", "").replace("$", "").replace("%", ""))
                continue
            except ValueError:
                pass
            label = cell_str
            break

        if not label:
            continue

        # Clean up: remove asterisks anywhere in the label, tidy whitespace
        label = label.replace("*", "").strip()
        label = re.sub(r"\s+", " ", label)  # collapse multiple spaces

        # Skip footer / footnote rows
        if _FOOTER_PATTERN.match(label):
            continue

        # Variance %
        varpct_val = None
        if col_idx["variance_pct"] is not None and col_idx["variance_pct"] < len(row):
            varpct_val = _to_number(row[col_idx["variance_pct"]])

        records.append({
            "Line Item":  label,
            "Budget":     budget_val,
            "Actual":     actual_val,
            "Variance %": varpct_val,
        })

    if not records:
        raise ValueError("No data rows found after the header row.")

    return pd.DataFrame(records)


def excel_to_text(file_bytes: bytes, max_rows: int = 150) -> str:
    """
    Converts the active sheet of an Excel file to a lean pipe-delimited text grid.
    Sent to the AI when the standard parser cannot read the file cleanly.

    To keep token count low:
    - Completely empty rows are skipped.
    - Columns that are entirely empty across all kept rows are dropped.
    - Output is capped at max_rows rows.
    This typically reduces a 20-column sparse finance file to 8-10 relevant columns.
    """
    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    # Read all rows up to max_rows, skipping completely empty ones
    raw_rows = []
    for row in ws.iter_rows(max_row=max_rows, values_only=True):
        if any(c is not None for c in row):
            raw_rows.append(row)

    if not raw_rows:
        return ""

    # Find which columns have at least one non-None value
    num_cols = max(len(r) for r in raw_rows)
    active_cols = [
        j for j in range(num_cols)
        if any(j < len(r) and r[j] is not None for r in raw_rows)
    ]

    # Build text using only active columns
    lines = []
    for row in raw_rows:
        cells = [str(row[j]) if j < len(row) and row[j] is not None else "" for j in active_cols]
        lines.append(" | ".join(cells))
    return "\n".join(lines)


def is_parse_suspicious(df: pd.DataFrame) -> bool:
    """
    Returns True if the standard pandas parse looks wrong — a signal to retry
    with AI extraction instead.

    A parse is suspicious when any of these are true:
    - More than 2 columns have unnamed/blank headers
    - The first column name is very long (a title row was mistaken for the header)
    - The first column is mostly empty (line item labels are in a different column)
    """
    cols = list(df.columns)
    unnamed_count = sum(
        1 for c in cols
        if str(c).startswith("Unnamed:") or str(c).strip() in ("", "None")
    )
    if unnamed_count > 2:
        return True
    first_col_name = str(cols[0]) if cols else ""
    if len(first_col_name) > 60:
        return True
    if len(cols) > 0 and df[cols[0]].isna().mean() > 0.7:
        return True
    return False


def dataframe_to_text(df: pd.DataFrame) -> str:
    """
    Converts a dataframe into a plain-text table string.
    This is what gets sent to the AI model — models read text, not Excel files.
    """
    return df.to_string(index=False)
