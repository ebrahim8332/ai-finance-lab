"""
Unified file parser for FinanceAI Lab.

Every module that accepts an Excel upload uses this parser.
It handles the two dominant real-world formats:

  VERTICAL   — line items as rows, Budget and Actual as columns.
               Most single-period management accounts look like this.

  HORIZONTAL — line items as rows, months/periods as columns.
               Common in monthly rolling reports and board packs.

How it works (two steps):

  Step 1 — Structure detection (AI, ~50 tokens)
    Only the first 8 rows are sent to the AI.
    The AI returns a JSON map: format, header row, column indices, periods.
    If the AI call fails, the robust openpyxl fallback takes over.

  Step 2 — Data extraction (openpyxl, zero tokens)
    openpyxl reads every data row using the exact column indices from step 1.
    Returns a clean 4-column dataframe every time:
      Line Item | Budget | Actual | Variance %

Usage in any module:
    from utils.file_parser import detect_structure, extract_data

    structure = detect_structure(file_bytes, chain)

    if structure["format"] == "horizontal":
        period = st.selectbox("Select period", structure["period_names"])
        df = extract_data(file_bytes, structure, period=period)
    else:
        df = extract_data(file_bytes, structure)
"""

import re
import json
import pandas as pd
from io import BytesIO


# ── Header row reader ─────────────────────────────────────────────────────

def _read_header_rows(file_bytes: bytes, n_rows: int = 8) -> str:
    """
    Reads the first n_rows of the active sheet and returns a compact text
    representation for the AI detection call.
    Only non-None cells are shown. Each line is: RowN: ColM=value | ColM=value
    Typically 150-300 tokens for an 8-row, 20-column header block.
    """
    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    lines = []
    for row_idx, row in enumerate(ws.iter_rows(max_row=n_rows, values_only=True)):
        cells = [
            (col_idx, str(val))
            for col_idx, val in enumerate(row)
            if val is not None
        ]
        if not cells:
            continue
        cell_str = " | ".join(f"Col{col}={val}" for col, val in cells)
        lines.append(f"Row{row_idx}: {cell_str}")

    return "\n".join(lines)


# ── AI structure detection ────────────────────────────────────────────────

_DETECT_SYSTEM = """You are a spreadsheet structure analyst for financial Excel files.
You will receive the first few rows of a file shown as "RowN: ColM=value | ColM=value".
Identify the file format and exact column indices. Return ONLY valid JSON — no explanation, no markdown.

VERTICAL file (Budget and Actual are columns, line items are rows):
{
  "format": "vertical",
  "header_row": 2,
  "label_col": 3,
  "budget_col": 11,
  "actual_col": 14,
  "variance_col": null,
  "variance_pct_col": 16
}

HORIZONTAL file (months or periods are columns, line items are rows):
{
  "format": "horizontal",
  "header_row": 0,
  "label_col": 0,
  "periods": [
    {"name": "Jan 2026", "budget_col": 1, "actual_col": 2, "variance_col": 3, "variance_pct_col": null},
    {"name": "Feb 2026", "budget_col": 4, "actual_col": 5, "variance_col": 6, "variance_pct_col": null},
    {"name": "YTD",      "budget_col": 13, "actual_col": 14, "variance_col": 15, "variance_pct_col": null}
  ]
}

Rules:
- header_row: 0-based row index where the actual column headers are (skip title/subtitle rows above)
- label_col: column index that contains line item names (often leftmost non-empty text column)
- All indices are 0-based integers
- Use null for any column not present
- For horizontal files: include ALL periods found — individual months AND any YTD / Full Year totals
- "Budget" may be labelled: Plan, Target, Approved, Forecast, Authorized, Annual
- "Actual" may be labelled: Actuals, Spend, Incurred, YTD Actual, Committed, Realized
- "Variance %" may be labelled: % Var, Var%, % to Budget, Delta %, Over/Under %
"""


def detect_structure(file_bytes: bytes, chain) -> dict:
    """
    Sends only the header rows to the AI and returns a structure dict.

    Returns a dict with at minimum:
      format        : "vertical" or "horizontal"
      header_row    : int
      label_col     : int
      budget_col    : int  (vertical) or None (horizontal — per period)
      actual_col    : int  (vertical) or None (horizontal — per period)
      variance_col  : int or None
      variance_pct_col : int or None
      periods       : list of period dicts (horizontal only)
      period_names  : list of period name strings (horizontal only)
      _raw          : the raw header text (kept for fallback)

    If the AI call fails, falls back to keyword-based detection.
    """
    header_text = _read_header_rows(file_bytes, n_rows=8)

    messages = [
        {"role": "system", "content": _DETECT_SYSTEM},
        {"role": "user",   "content": f"FILE HEADER ROWS:\n{header_text}"},
    ]

    try:
        response, _ = chain.complete(messages, timeout=30)
        clean = re.sub(r"```(?:json)?|```", "", response).strip()
        match = re.search(r"\{.*\}", clean, re.DOTALL)
        if not match:
            raise ValueError("No JSON object found in AI response.")
        structure = json.loads(match.group(0))

        # Add convenience period_names list for horizontal files
        if structure.get("format") == "horizontal":
            structure["period_names"] = [p["name"] for p in structure.get("periods", [])]
        else:
            structure["period_names"] = []

        structure["_raw"] = header_text
        structure["_source"] = "ai"
        structure = _validate_structure(structure, file_bytes)
        return structure

    except Exception:
        # AI failed — fall back to keyword-based detection
        return _detect_structure_fallback(file_bytes, header_text)


def _validate_structure(structure: dict, file_bytes: bytes) -> dict:
    """
    Validates the AI-detected column indices against the actual file.
    Corrects off-by-one errors (very common with AI column counting) by checking
    adjacent columns when the detected one doesn't contain the expected keyword.
    Also validates the header_row index — if the detected row is empty, steps forward.
    """
    import openpyxl

    BUDGET_KW  = ["budget", "plan", "target", "bud", "approved", "authorized"]
    ACTUAL_KW  = ["actual", "actuals", "real", "result", "spend", "incurred"]
    VARPCT_KW  = ["% var", "var %", "variance %", "% variance", "% to budget", "delta %"]

    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(max_row=15, values_only=True))

    # Validate header_row — if the detected row is mostly empty, step forward
    header_idx = structure.get("header_row", 0)
    for offset in range(3):
        candidate = header_idx + offset
        if candidate >= len(all_rows):
            break
        row = all_rows[candidate]
        non_null = [c for c in row if c is not None]
        if len(non_null) >= 3:
            if offset > 0:
                structure["header_row"] = candidate
            break

    header = all_rows[structure["header_row"]] if structure["header_row"] < len(all_rows) else ()

    def best_col(col_idx, keywords):
        """Return col_idx if it matches keywords, else search ±3 adjacent columns."""
        if col_idx is None:
            return col_idx
        # Check detected column first
        if col_idx < len(header) and header[col_idx] is not None:
            if any(k in str(header[col_idx]).lower() for k in keywords):
                return col_idx
        # Try adjacent columns within a window of 3
        for offset in [1, -1, 2, -2, 3, -3]:
            adj = col_idx + offset
            if 0 <= adj < len(header) and header[adj] is not None:
                if any(k in str(header[adj]).lower() for k in keywords):
                    return adj
        return col_idx  # return original if no better match found

    if structure.get("format") == "vertical":
        structure["budget_col"]       = best_col(structure.get("budget_col"), BUDGET_KW)
        structure["actual_col"]       = best_col(structure.get("actual_col"), ACTUAL_KW)
        structure["variance_pct_col"] = best_col(structure.get("variance_pct_col"), VARPCT_KW)
    else:
        # Validate each period's columns
        for period in structure.get("periods", []):
            period["budget_col"] = best_col(period.get("budget_col"), BUDGET_KW)
            period["actual_col"] = best_col(period.get("actual_col"), ACTUAL_KW)

    return structure


def _detect_structure_fallback(file_bytes: bytes, header_text: str) -> dict:
    """
    Keyword-based structure detection used when the AI call fails.
    Handles common column names only — less flexible than the AI path.
    """
    import openpyxl

    BUDGET_KW      = ["budget", "plan", "target", "bud", "approved", "authorized"]
    ACTUAL_KW      = ["actual", "actuals", "real", "result", "spend", "incurred"]
    VARPCT_KW      = ["% var", "var %", "variance %", "% variance", "pct var",
                      "var to budget", "% to budget", "delta %"]
    VAR_KW         = ["variance", "var $", "var(", "diff", "over/under", "delta"]
    MONTH_KEYWORDS = ["jan", "feb", "mar", "apr", "may", "jun",
                      "jul", "aug", "sep", "oct", "nov", "dec",
                      "q1", "q2", "q3", "q4", "ytd", "full year", "total"]

    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(max_row=8, values_only=True))

    # Find header row
    header_row_idx = 0
    for i, row in enumerate(all_rows):
        non_null = [c for c in row if c is not None]
        if len(non_null) >= 3:
            for cell in non_null:
                if isinstance(cell, str):
                    low = cell.lower()
                    if any(k in low for k in BUDGET_KW + ACTUAL_KW):
                        header_row_idx = i
                        break
            else:
                continue
            break

    header = all_rows[header_row_idx] if header_row_idx < len(all_rows) else ()

    # Detect if horizontal by checking if month keywords appear in header
    month_cols = [
        i for i, c in enumerate(header)
        if c and any(k in str(c).lower() for k in MONTH_KEYWORDS)
    ]
    is_horizontal = len(month_cols) >= 3

    if is_horizontal:
        # Build period list: look for pairs of budget/actual columns under month headers
        periods = []
        # Simplified: treat each month column as an actual-only period
        for col_idx in month_cols:
            periods.append({
                "name": str(header[col_idx]),
                "budget_col": None,
                "actual_col": col_idx,
                "variance_col": None,
                "variance_pct_col": None,
            })
        return {
            "format": "horizontal",
            "header_row": header_row_idx,
            "label_col": 0,
            "budget_col": None,
            "actual_col": None,
            "variance_col": None,
            "variance_pct_col": None,
            "periods": periods,
            "period_names": [p["name"] for p in periods],
            "_raw": header_text,
            "_source": "fallback",
        }

    # Vertical — find columns by keyword
    budget_col = actual_col = variance_col = variance_pct_col = label_col = None
    for i, cell in enumerate(header):
        if cell is None:
            continue
        low = str(cell).lower().strip()
        if budget_col      is None and any(k in low for k in BUDGET_KW):
            budget_col = i
        elif actual_col    is None and any(k in low for k in ACTUAL_KW):
            actual_col = i
        elif variance_pct_col is None and any(k in low for k in VARPCT_KW):
            variance_pct_col = i
        elif variance_col  is None and any(k in low for k in VAR_KW):
            variance_col = i

    # Label column: first column with text data in the first data row
    first_data_row = all_rows[header_row_idx + 1] if header_row_idx + 1 < len(all_rows) else ()
    for i, cell in enumerate(first_data_row):
        if cell and isinstance(cell, str) and len(cell.strip()) > 1:
            label_col = i
            break
    if label_col is None:
        label_col = 0

    return {
        "format": "vertical",
        "header_row": header_row_idx,
        "label_col": label_col,
        "budget_col": budget_col,
        "actual_col": actual_col,
        "variance_col": variance_col,
        "variance_pct_col": variance_pct_col,
        "periods": [],
        "period_names": [],
        "_raw": header_text,
        "_source": "fallback",
    }


# ── Data extraction ───────────────────────────────────────────────────────

_FOOTER_RE = re.compile(
    r"^(source|note|footnote|all figures|simulated|prepared by|disclaimer)",
    re.IGNORECASE,
)


def _to_number(val) -> float | None:
    """Converts a cell value to float. Returns None if not numeric."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).replace(",", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return None


def _get_cell(row: tuple, col_idx) -> object:
    """Safe cell read — returns None if col_idx is None or out of range."""
    if col_idx is None or col_idx >= len(row):
        return None
    return row[col_idx]


def extract_data(file_bytes: bytes, structure: dict, period: str | None = None) -> pd.DataFrame:
    """
    Extracts data from the Excel file using column indices from detect_structure().

    For vertical files: period is ignored.
    For horizontal files: period must match one of structure["period_names"].

    Returns a clean dataframe with columns: Line Item, Budget, Actual, Variance %
    Raises ValueError if extraction fails.
    """
    import openpyxl

    fmt = structure.get("format", "vertical")

    # Resolve column indices for the chosen period (horizontal) or direct (vertical)
    if fmt == "horizontal":
        if not period:
            raise ValueError("Period must be specified for horizontal files.")
        period_defs = structure.get("periods", [])
        chosen = next((p for p in period_defs if p["name"] == period), None)
        if chosen is None:
            raise ValueError(f"Period '{period}' not found in file structure.")
        budget_col      = chosen.get("budget_col")
        actual_col      = chosen.get("actual_col")
        variance_col    = chosen.get("variance_col")
        variance_pct_col = chosen.get("variance_pct_col")
    else:
        budget_col      = structure.get("budget_col")
        actual_col      = structure.get("actual_col")
        variance_col    = structure.get("variance_col")
        variance_pct_col = structure.get("variance_pct_col")

    label_col   = structure.get("label_col", 0)
    header_row  = structure.get("header_row", 0)

    if budget_col is None and actual_col is None:
        raise ValueError(
            "Could not identify Budget and Actual columns. "
            "Ensure the file has columns labelled 'Budget' and 'Actual' (or similar)."
        )

    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))

    records = []
    for row in all_rows[header_row + 1:]:
        if not row or all(c is None for c in row):
            continue

        budget_val  = _to_number(_get_cell(row, budget_col))
        actual_val  = _to_number(_get_cell(row, actual_col))

        # Skip rows with no numeric data — section headers like "FACILITY EXPENSES"
        if budget_val is None and actual_val is None:
            continue

        # Find label: first non-empty text cell scanning left to right
        label = None
        for cell in row:
            if cell is None:
                continue
            cell_str = str(cell).strip()
            if not cell_str or cell_str.lower() in ("none", "nan"):
                continue
            try:
                float(cell_str.replace(",", "").replace("$", "").replace("%", ""))
                continue  # it's a number, skip
            except ValueError:
                pass
            label = cell_str
            break

        if not label:
            continue

        # Clean label: remove asterisks, collapse spaces
        label = label.replace("*", "").strip()
        label = re.sub(r"\s+", " ", label)

        if _FOOTER_RE.match(label):
            continue

        # Variance %: prefer explicit column, otherwise compute from budget/actual
        var_pct = _to_number(_get_cell(row, variance_pct_col))
        if var_pct is None and variance_col is not None:
            var_val = _to_number(_get_cell(row, variance_col))
            if var_val is not None and budget_val:
                var_pct = var_val / budget_val

        records.append({
            "Line Item":  label,
            "Budget":     budget_val,
            "Actual":     actual_val,
            "Variance %": var_pct,
        })

    if not records:
        raise ValueError("No data rows found. Check that the file has Budget and Actual columns with numbers.")

    return pd.DataFrame(records)
